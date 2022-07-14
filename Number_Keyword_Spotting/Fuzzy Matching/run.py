from asyncio import constants
import numbers
import get_nums
import os
from num2words import num2words 
from num_to_words.utils import constants
import csv
import openpyxl
from pathlib import Path





#new code begins

from libindic.utils import servicemethod
from libindic.soundex import Soundex


_all_ = ['InexactSearch', 'getInstance']


class InexactSearch(object):
    """
       This class provides methods for fuzzy searching using word
       distance as well as phonetics.
    """

    def __init__(self):
        self.sx = Soundex()

    def _countCommon(self, shrtBigr, lngBigr, average):
        common = 0.0
        for indexShrt, bigr in enumerate(shrtBigr):
            if bigr in lngBigr:
                indexLng = lngBigr.index(bigr)
                if indexLng == indexShrt:
                    common += 1.0
                else:
                    dislocation = (indexLng - indexShrt) / average
                    if dislocation < 0:
                        dislocation *= -1
                    common += 1.0 - dislocation

        return common

    def _createBigram(self, string):
        bigram = []
        for i in range(1, len(string)):
            bigram.append(string[i - 1:i + 1])

        return bigram

    def bigram_average(self, str1, str2):
        """Return approximate string comparator measure (between 0.0 and 1.0)
        using bigrams.

        :param str1: string 1 for comparison
        :str1 type : str
        :param str2: string 2 for comparison
        :str2 type : str
        :returns: int score between 0.0 and 1.0

        >>> score = bigram_avearage(str1, str2)
        0.7


        Bigrams are two-character sub-strings contained in a
        string. For example, 'peter' contains the bigrams:
        pe,et,te,er.

        This routine counts the number of common bigrams and divides
        by the average number of bigrams. The resulting number is
        returned.
        """

        if str1 == str2:
            return 1

        bigr1 = self._createBigram(str1)
        bigr2 = self._createBigram(str2)

        average = (len(bigr1) + len(bigr2)) / 2.0

        common = 0.0

        if len(bigr1) < len(bigr2):  # Count using the shorter bigram list
            common = self._countCommon(bigr1, bigr2, average)
        else:
            common = self._countCommon(bigr2, bigr1, average)

        return common / average

    def compare(self, string1, string2):
        ''' Compare strings using soundex if not possible gives
        biggram avearage.

        :param str1: string 1 for comparison.
        :type str1: str.
        :param str2: string 2 for comparison
        :type str2: str.
        :returns: int score between 0.0 and 1.0

        '''
        weight = 0
        if string1 == string2:
            return 1.0

        soundex_match = self.sx.compare(string1, string2)

        if soundex_match == 1:
            weight = 0.9

        if soundex_match == 2:
            weight = 0.8

        if weight == 0:
            return self.bigram_average(string1, string2)

        return weight

    @servicemethod
    def search(self, text, key):
        '''Searches for the key in the given text. This function uses
        :method: `InexactSearch.compare` for doing approx search.

        :param text: text in which search has to be done.
        :type text: str.
        :param key: key which has to be searched
        :type key: str.
        :returns: A dictionary with words in the string as keys and
        the score against the key as the value
        '''
        key = key.strip()
        words = text.split()
        search_results = {}
        for word in words:
            word = word.strip()
            search_results[word] = self.compare(word, key)

        return search_results


def getInstance():
    '''This function returns instance of :py:class:`~InexactSearch`
    '''
    return InexactSearch()

ob=getInstance()



#new code ends

en_xlsx_file=Path("data","en_asr_hitachi.xlsx")
en_asr = openpyxl.load_workbook(en_xlsx_file) 

hi_xlsx_file=Path('data','hi_asr_hitachi.xlsx')
hi_asr = openpyxl.load_workbook(hi_xlsx_file) 

bn_xlsx_file=Path('data','bn_asr_hitachi.xlsx')
bn_asr = openpyxl.load_workbook(bn_xlsx_file) 

ta_xlsx_file=Path('data','ta_asr_hitachi.xlsx')
ta_asr = openpyxl.load_workbook(ta_xlsx_file) 


number=dict()
powers=dict()

#ENGLISH

number["en"]={"zero":0, "one":1, "two":2, "three":3,"four":4,"five":5, "six":6, "seven":7, "eight":8, "nine":9,"ten":10, "eleven":11, "twelve":12,
    "thirteen":13, "fourteen":14, "fifteen":15, "sixteen":16, "seventeen":17, "eighteen":18, "nineteen":19, "twenty":20,
    "thirty":30, "fourty":40, "fifty":50, "sixty":60,"seventy":70, "eighty":80, "ninety":90,}
powers["en"]={"hundred":100, "thousand":1000, "lakh":100000, "crore":10000000, "million":1000000, "billion":1000000000,'and':'and'}




#HINDI

number["hi"]=constants.NUM_DICT["hi"]
powers["hi"]=dict()
powers["hi"][number["hi"].pop("100")]="100"
powers["hi"][number["hi"].pop("1000")]="1000"
powers["hi"][number["hi"].pop("100000")]="100000"
powers["hi"][number["hi"].pop("10000000")]="10000000"
powers["hi"][number["hi"].pop("1000000000")]="1000000000"

number["hi"]= dict((v,k) for k,v in number["hi"].items())


#TAMIL

number["ta"]=constants.NUM_DICT["ta"]
powers["ta"]=dict()
powers["ta"][number["ta"].pop("100")]="100"
powers["ta"][number["ta"].pop("1000")]="1000"
powers["ta"][number["ta"].pop("100000")]="100000"
powers["ta"][number["ta"].pop("10000000")]="10000000"
powers["ta"][number["ta"].pop("1000000000")]="1000000000"

number["ta"]= dict((v,k) for k,v in number["ta"].items())


#BENGALI

number["bn"]=constants.NUM_DICT["bn"]
powers["bn"]=dict()
powers["bn"][number["bn"].pop("100")]="100"
powers["bn"][number["bn"].pop("1000")]="1000"
powers["bn"][number["bn"].pop("100000")]="100000"
powers["bn"][number["bn"].pop("10000000")]="10000000"
powers["bn"][number["bn"].pop("1000000000")]="1000000000"

number["bn"]= dict((v,k) for k,v in number["bn"].items())

# cwd = os.getcwd()
# directory="data"
# path=os.path.join(cwd,directory)
# try: 
#     os.mkdir(path) 
# except OSError as error: 
#     print("Directory already exists")  
# f = open(f'data/numbers_en.txt', "w")
# for x in range(100000):
#     temp=num2words(x,lang='en').replace(',','')
#     temp=temp.replace('-',' ')
#     f.write(temp.replace(' and ',' ')+" "+ str(x)+"\n")



print("Language Code= English: en     Tamil: ta     Hindi: hi       Bengali: bn")
language=input("Enter language code:")

if language=="en":
    active_sheet= en_asr.active
elif language=="hi":
    active_sheet=hi_asr.active
elif language=="bn":
    active_sheet=bn_asr.active
else:
    active_sheet=ta_asr.active

# sentence = input("Enter the string:")
# sentence=sentence.replace('.','')
# sentence=sentence.replace('-',' ')
# sentence=sentence.lower()
# sentence_list=sentence.split()

col_count=0
for col in active_sheet.iter_cols(max_col=1,max_row=11):
    for cell in col:
        col_count=col_count+1
        if col_count== 1:
            language=cell.value
            keywords=(list)(number[language].keys())
            keywords.extend((list)(powers[language].keys()))
            continue
        sentence=cell.value
        sentence=sentence.replace(',','')
        sentence=sentence.replace('.','')
        sentence=sentence.replace('-',' ')
        sentence=sentence.lower()
        sentence_list=sentence.split()

        counter=0
        for x in range(len(sentence_list)):
            for keyword in keywords:
                score=ob.bigram_average(sentence_list[x],(keyword))
                # if(number[language].get(sentence_list[x])!=None or powers[language].get(sentence_list[x])!=None):
                if(score>0.7):
                    # if( counter==0 and powers[language].get(sentence_list[x])=='and'):
                    #     break
                    # if(counter==0):
                    #     counter=(counter+1)%2
                    #     sentence_list[x]="{"+sentence_list[x]
                    # if(counter==1):
                    #     if(x==len(sentence_list)-1):
                    #         sentence_list[x]=sentence_list[x]+"}"
                    #         counter=(counter+1)%2
                    if(counter==0):
                        sentence_list[x]="{"+sentence_list[x]
                        counter=1
                        break
                    if(counter==1 and x==len(sentence_list)-1):
                        sentence_list[x]=sentence_list[x]+"}"
                    if(counter==1):
                        break

                    # print("Found ",keyword)
                    # print("At pos",x)
                    # print("Score=",score)
                else:
                    if(counter==1 and keywords.index(keyword)==len(keywords)-1):
                        sentence_list[x-1]=sentence_list[x-1]+"}"
                        counter=0
                    continue

        final_sentence=""
        for x in range(len(sentence_list)):
            final_sentence=final_sentence+" " + sentence_list[x]
        active_sheet.cell(row=col_count,column=2).value=final_sentence
        print(final_sentence)

en_asr.save("data\\en_asr_hitachi.xlsx")
hi_asr.save("data\\hi_asr_hitachi.xlsx")
bn_asr.save("data\\bn_asr_hitachi.xlsx")
ta_asr.save("data\\ta_asr_hitachi.xlsx")











